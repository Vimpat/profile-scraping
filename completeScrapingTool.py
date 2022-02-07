from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import openpyxl
import requests  # to get image from the web
import shutil  # to save it locally

# Add the username and password for the LinkedIn and Facebook profiles
USERNAME_LINKEDIN = ""
PASSWORD_LINKEDIN = ""
USERNAME_FACEBOOK = ""
PASSWORD_FACEBOOK = ""

#EXAMPLE:
# USERNAME_LINKEDIN = "testli"
# PASSWORD_LINKEDIN = "123li"
# USERNAME_FACEBOOK = "testfb"
# PASSWORD_FACEBOOK = "123fb"


# logins into linkedIn
def login_linkedin():
    driver.get("https://www.linkedin.com/login")
    time.sleep(3)
    email = driver.find_element(By.ID, "username")
    email.send_keys(USERNAME_LINKEDIN)
    password = driver.find_element(By.ID, "password")
    password.send_keys(PASSWORD_LINKEDIN)
    password.send_keys(Keys.RETURN)
    time.sleep(2)


# Logins into facebook
def login_facebook():
    driver.get('https://www.facebook.com/login')
    time.sleep(3)
    driver.find_element(By.CLASS_NAME, "_42ft._4jy0._9xo7._4jy3._4jy1.selected._51sy").click()
    email = driver.find_element(By.ID, "email")
    email.send_keys(USERNAME_FACEBOOK)
    password = driver.find_element(By.ID, "pass")
    password.send_keys(PASSWORD_FACEBOOK)
    password.send_keys(Keys.RETURN)
    time.sleep(2)


# downloads the image from the link
# @image_url - the link containing the image to download
# @name - the name of the person whose photo is downloaded
def download_image(image_url, person_name):
    filename = person_name + ".jpg"
    filename = filename.replace(" ", "_")
    r = requests.get(image_url, stream=True)
    # Check if the image was retrieved successfully
    if r.status_code == 200:
        # Set decode_content value to True, otherwise the downloaded image file's size will be zero.
        r.raw.decode_content = True

        # Open a local file with wb ( write binary ) permission.
        with open("photos/" + filename, 'wb') as f:
            shutil.copyfileobj(r.raw, f)
    else:
        print('Image Couldn\'t be retrieved ' + filename)


# Returns the image from an image link
# @page - the link of the picture to be downloaded
def get_image_complete_link(page):
    driver.get(page)
    time.sleep(3)
    image = driver.find_element(By.TAG_NAME, "img")
    return image.get_attribute("src")


# Returns the profile picture of a facebook profile
# @page - the link of the facebook profile
def facebook_get_profile_picture(page):
    driver.get(page)
    time.sleep(3)
    anchors = driver.find_elements(By.TAG_NAME, 'a')

    for anchor in anchors:
        if anchor.get_attribute('aria-label') == "Link to open profile cover photo" or \
                anchor.get_attribute('aria-label') == "Link to open page cover photo":
            anchors.remove(anchor)

    anchors = [a.get_attribute('href') for a in anchors]
    anchors = [a for a in anchors if str(a).startswith("https://www.facebook.com/photo")]

    return get_image_complete_link(anchors[0])


s = Service(r'chromedriver_win32/chromedriver.exe')
chrome_options = Options()
chrome_options.add_argument("--headless")
driver = webdriver.Chrome(service=s, options=chrome_options)

# Login into facebook and linkedin
login_linkedin()
login_facebook()

peopleDb = openpyxl.load_workbook('Database.xlsx')
peopleSheet = peopleDb.active

for j in range(1, 12):
    i = j * 10
    name = peopleSheet.cell(row=i, column=1).value + " " + peopleSheet.cell(row=i, column=2).value
    photo = ""
    photoLink = peopleSheet.cell(row=i, column=3).value
    print(photoLink)
    try:
        if 'linkedin' in photoLink:
            if 'detail/photo' in photoLink:
                photoLink = photoLink.replace("detail/photo/", "overlay/photo/")
            elif 'overlay/photo/' not in photoLink:
                photoLink = photoLink + 'overlay/photo'
            photo = get_image_complete_link(photoLink)

        elif 'facebook' in photoLink:
            if '/photo' in photoLink:
                photo = get_image_complete_link(photoLink)
            else:
                photo = facebook_get_profile_picture(photoLink)

        elif 'twitter' in photoLink:
            if '/photo' not in photoLink:
                photoLink = photoLink + '/photo'
            photo = get_image_complete_link(photoLink)
        else:
            photo = photoLink
        download_image(photo, name)
    except:
        print("There is no image link for " + name)
driver.quit()
